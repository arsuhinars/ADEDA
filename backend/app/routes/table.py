from io import BytesIO
from fastapi import APIRouter, File, Depends
from fastapi.responses import StreamingResponse
from pydantic import ValidationError
from openpyxl import load_workbook, Workbook

from ..internal.errors import ParseError
from ..schemas import HouseBrief, \
    HouseSegment, HouseMaterial, HouseState, ExportTable
from ..dependencies import check_authorization

router = APIRouter(
    prefix='/table',
    tags=['table']
)

@router.post(
    '/parse',
    response_model=list[HouseBrief],
    response_model_exclude_none=True,
    dependencies=[Depends(check_authorization)],
    summary='Выполнить парсинг таблицы с квартирами'
    )
def parse_house_table(file: bytes = File()):
    """
    Данный метод выполняем парсинг исходной Excel таблицы со всеми эталонными
    квартирами. В теле запроса необходимо передать сам файл таблицы.
    """
    with BytesIO(file) as stream:
        try:
            wb = load_workbook(stream, read_only=True)
        except Exception as err:
            raise ParseError('Unable to load table')
        ws = wb[wb.sheetnames[0]]

        houses: list[HouseBrief] = []

        for row in ws.iter_rows(2, max_col=11):
            # Сегмент дома
            if type(row[2].value) is not str:
                raise ParseError(
                    f'Unable to parse house segment at {row[2].coordinate}'
                )

            match row[2].value.strip().lower():
                case 'новостройка':
                    segment = HouseSegment.NEW
                case 'современное жилье':
                    segment = HouseSegment.MODERN
                case 'старый жилой фонд':
                    segment = HouseSegment.OLD
                case _:
                    raise ParseError(
                        f'Unable to parse house segment at {row[2].coordinate}'
                    )
            
            # Материал стен
            if type(row[4].value) is not str:
                raise ParseError(
                    f'Unable to parse wall material at {row[4].coordinate}'
                )
            
            match row[4].value.strip().lower():
                case 'кирпич':
                    material = HouseMaterial.BRICK
                case 'панель':
                    material = HouseMaterial.PANEL
                case 'монолит':
                    material = HouseMaterial.MONOLIT
                case _:
                    raise ParseError(
                        f'Unable to parse wall material at {row[4].coordinate}'
                    )
            
            # Наличие балкона/лоджии
            if type(row[8].value) is not str:
                raise ParseError(
                    f'Unable to parse balcony at {row[8].coordinate}'
                )
            
            match row[8].value.strip().lower():
                case 'да':
                    has_balcony = True
                case 'нет':
                    has_balcony = False
                case _:
                    raise ParseError(
                        f'Unable to parse balcony at {row[8].coordinate}'
                    )
            
            # Состояние дома
            if type(row[10].value) is not str:
                raise ParseError(
                    f'Unable to parse house state at {row[10].coordinate}'
                )
            
            match row[10].value.strip().lower():
                case 'без отделки':
                    state = HouseState.NO_DECORATION
                case 'муниципальный ремонт':
                    state = HouseState.STATE_DECORATION
                case 'современная отделка':
                    state = HouseState.MODERN_DECORATION
                case _:
                    raise ParseError(
                        f'Unable to parse house state at {row[10].coordinate}'
                    )

            # Кол-во комнат
            if type(row[1].value) is str and row[1].value.strip().lower() == 'студия':
                rooms_count = 1
            else:
                try:
                    rooms_count = int(row[1].value)
                except ValueError:
                    raise ParseError(
                        f'Unable to parse rooms count at {row[1].coordinate}'
                    )

            try:
                houses.append(HouseBrief(
                    location=row[0].value,
                    rooms_count=rooms_count,
                    segment=segment,
                    floors_count=row[3].value,
                    material=material,
                    floor=row[5].value,
                    flat_area=row[6].value,
                    kitchen_area=row[7].value,
                    has_balcony=has_balcony,
                    metro_distance=row[9].value,
                    state=state
                ))
            except ValidationError as err:
                raise ParseError(str(err))
    return houses


@router.post(
    '/create',
    dependencies=[Depends(check_authorization)],
    summary='Экспортировать таблицу в Excel'
    )
def create_house_table(export_table: ExportTable):
    """
    Метод выполняет экспорт полученных аналогов в Excel таблицу.

    Поля в теле запроса:
    * `reference` - объект эталонной квартиры
    * `analog` - массив квартир-аналогов, где поле `house` - объект квартиры,
    `adjustments` - объект, содержащий все примененнные корректировки.
    """
    ref_house = export_table.reference
    analogs = export_table.analogs

    wb = Workbook()
    ws = wb.active
    ws.title = 'Расчет'

    cell_coord = lambda r, c: ws.cell(r, c).coordinate

    def apply_style(start: str, end: str, style):
        cell_range = ws[start:end]
        for row in cell_range:
            for cell in row:
                cell.style = style

    ws['A1'] = 'Элементы сравнения'
    ws['A2'] = 'Источник информации'
    ws['A3'] = 'Площадь, кв.м.'
    ws['A4'] = 'Этаж расположения/этажность'
    ws['A5'] = 'Адрес'
    ws['A6'] = 'Удаленность от метро, мин. пешком'
    ws['A7'] = 'Комнатность'
    ws['A8'] = 'Материал стен'
    ws['A9'] = 'Сегмент'
    ws['A10'] = 'Площадь кухни'
    ws['A11'] = 'Наличие балкона/лоджии'
    ws['A12'] = 'Состояние отделки'
    ws['A13'] = 'Цена предложения, руб.'
    ws.merge_cells('A13:B13')
    ws['A14'] = 'Цена предложения, руб./кв.м'
    ws.merge_cells('A14:B14')

    ws['A15'] = 'Корректировка на торг'
    ws.merge_cells('A15:B16')
    ws['A17'] = 'Корректировка на площадь'
    ws.merge_cells('A17:B18')
    ws['A19'] = 'Корректировка на удаленность от метро'
    ws.merge_cells('A19:B20')
    ws['A21'] = 'Корректировка на этаж'
    ws.merge_cells('A21:B22')
    ws['A23'] = 'Корректировка на площадь кухни'
    ws.merge_cells('A23:B24')
    ws['A25'] = 'Корректировка на наличие балкона/лоджии'
    ws.merge_cells('A25:B26')
    ws['A27'] = 'Корректировка на ремонт'
    ws.merge_cells('A27:B29')
    
    ws['A30'] = 'Размер примененных корректировок %'
    ws.merge_cells('A30:B30')
    ws['A31'] = 'Вес аналога'
    ws.merge_cells('A31:B31')
    ws['A32'] = 'Рыночная стоимость, руб./кв.м.'
    ws.merge_cells('A32:B32')
    ws['A33'] = 'Рыночная стоимость, руб.'
    ws.merge_cells('A33:B33')


    ws['B1'] = 'Объект оценки'
    ws['B2'] = '-'
    ws['B3'] = ref_house.flat_area
    ws['B4'] = f'{ref_house.floor} из {ref_house.floors_count}'
    ws['B5'] = ref_house.location
    ws['B6'] = ref_house.metro_distance
    ws['B7'] = ref_house.rooms_count
    ws['B8'] = ref_house.material.get_verbose_value()
    ws['B9'] = ref_house.segment.get_verbose_value()
    ws['B10'] = ref_house.kitchen_area
    ws['B11'] = 'Да' if ref_house.has_balcony else 'Нет'
    ws['B12'] = ref_house.state.get_verbose_value()

    weight_div = '+'.join(map(
        lambda col: f'1/{cell_coord(30, col)}',
        range(3, 3 + len(analogs))
    ))

    for i in range(len(analogs)):
        col = i + 3

        house = analogs[i].house
        adjustments = analogs[i].adjustments

        ws.cell(1, col).value = f'Аналог {i + 1}'

        ws.cell(2, col).value = house.source_service.get_verbose_value()
        ws.cell(2, col).hyperlink = house.url
        ws.cell(2, col).style = 'Hyperlink'

        ws.cell(3, col).value = house.flat_area
        ws.cell(4, col).value = f'{house.floor} из {house.floors_count}'
        ws.cell(5, col).value = house.location
        ws.cell(6, col).value = house.metro_distance
        ws.cell(7, col).value = house.rooms_count
        ws.cell(8, col).value = house.material.get_verbose_value()
        ws.cell(9, col).value = house.segment.get_verbose_value()
        ws.cell(10, col).value = house.kitchen_area
        ws.cell(11, col).value = 'Да' if house.has_balcony else 'Нет'
        ws.cell(12, col).value = house.state.get_verbose_value()
        ws.cell(13, col).value = house.price
        ws.cell(13, col).style = 'Comma [0]'
        ws.cell(14, col).value = f'={cell_coord(13, col)}/{cell_coord(3, col)}'
        ws.cell(14, col).style = 'Comma [0]'

    
        formula = lambda p: f'={cell_coord(p-1, col)}*(1 + {cell_coord(p, col)})'

        ws.cell(15, col).value = adjustments.trade / 100
        ws.cell(15, col).style = 'Percent'
        ws.cell(16, col).value = formula(15)
        ws.cell(16, col).style = 'Comma [0]'

        ws.cell(17, col).value = adjustments.area / 100
        ws.cell(17, col).style = 'Percent'
        ws.cell(18, col).value = formula(17)
        ws.cell(18, col).style = 'Comma [0]'

        ws.cell(19, col).value = adjustments.metro / 100
        ws.cell(19, col).style = 'Percent'
        ws.cell(20, col).value = formula(19)
        ws.cell(20, col).style = 'Comma [0]'

        ws.cell(21, col).value = adjustments.floor / 100
        ws.cell(21, col).style = 'Percent'
        ws.cell(22, col).value = formula(21)
        ws.cell(22, col).style = 'Comma [0]'

        ws.cell(23, col).value = adjustments.kitchen_area / 100
        ws.cell(23, col).style = 'Percent'
        ws.cell(24, col).value = formula(23)
        ws.cell(24, col).style = 'Comma [0]'

        ws.cell(25, col).value = adjustments.balcony / 100
        ws.cell(25, col).style = 'Percent'
        ws.cell(26, col).value = formula(25)
        ws.cell(26, col).style = 'Comma [0]'

        ws.cell(27, col).value = f'={cell_coord(28, col)}/{cell_coord(26, col)}'
        ws.cell(27, col).style = 'Percent'
        ws.cell(28, col).value = adjustments.repairs
        ws.cell(29, col).value = formula(27)
        ws.cell(29, col).style = 'Comma [0]'

        ws.cell(30, col).value = \
            f'=100*(ABS({cell_coord(15, col)}) + ABS({cell_coord(17, col)})' \
            f' + ABS({cell_coord(19, col)})' \
            f' + ABS({cell_coord(21, col)}) + ABS({cell_coord(23, col)})' \
            f' + ABS({cell_coord(25, col)}) + ABS({cell_coord(27, col)}))'
        
        ws.cell(31, col).value = f'=(1/{cell_coord(30, col)})/({weight_div})'
    
    ws.cell(32, 3).value = \
        f'=ROUND(SUMPRODUCT({cell_coord(29, 3)}:{cell_coord(29, 2 + len(analogs))},' \
        f'{cell_coord(31, 3)}:{cell_coord(31, 2 + len(analogs))}),-1)'
    ws.cell(32, 3).style = 'Comma [0]'
    ws.cell(33, 3).value = f'={cell_coord(32, 3)}*B3'
    ws.cell(33, 3).style = 'Comma [0]'

    apply_style('A1', cell_coord(1, 2 + len(analogs)), 'Accent1')
    apply_style('A2', 'A12', '40 % - Accent1')
    apply_style('A13', 'B14', '40 % - Accent1')
    apply_style('A15', 'B33', '40 % - Accent1')

    apply_style('C32', 'C33', 'Comma [0]')

    stream = BytesIO()
    wb.save(stream)
    wb.close()

    stream.seek(0)

    return StreamingResponse(
        content=stream,
        headers={
            'Content-Disposition': 'attachment; filename="output.xlsx"'
        }
    )
